import csv
import io
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from supabase import Client

from dependencies import (
    CurrentUser,
    get_service_client,
    require_super_admin,
    require_super_admin_or_manager,
)

router = APIRouter(prefix="/exports", tags=["exports"])

COLUMNS = {
    "properties": [
        "id", "title", "property_type", "bedrooms", "bathrooms",
        "rent_amount", "rent_period", "state", "city", "area",
        "address", "status", "manager_phone", "created_at",
    ],
    "tenants": [
        "id", "first_name", "last_name", "email", "phone",
        "status", "created_at",
    ],
    "leases": [
        "id", "property_id", "tenant_id", "monthly_rent",
        "start_date", "end_date", "status", "created_at",
    ],
    "payments": [
        "id", "lease_id", "tenant_id", "amount", "status",
        "payment_type", "due_date", "paid_date", "notes", "created_at",
    ],
    "boosts": [
        "id", "property_id", "manager_id", "amount_paid",
        "duration_days", "started_at", "expires_at", "status",
        "transaction_id", "payment_method", "created_at",
    ],
    "managers": [
        "id", "user_id", "email", "full_name", "phone",
        "role", "status", "created_at",
    ],
}

TABLE_MAP = {
    "properties": "properties",
    "tenants": "tenants",
    "leases": "leases",
    "payments": "payments",
    "boosts": "property_boosts",
    "managers": "profiles",
}


def _build_query(
    supabase: Client,
    resource: str,
    current_user: CurrentUser,
    status_filter: str | None = None,
    state_filter: str | None = None,
    property_type_filter: str | None = None,
    skip: int = 0,
    limit: int = 10000,
):
    table = TABLE_MAP[resource]
    query = supabase.table(table).select(",".join(COLUMNS[resource]))

    if resource == "managers":
        query = query.eq("role", "house_manager")
    elif current_user.role == "house_manager" and resource != "boosts":
        if resource == "properties":
            query = query.eq("owner_id", current_user.id)
        elif resource == "tenants":
            query = query.eq("owner_id", current_user.id)
        elif resource == "leases":
            query = query.eq("owner_id", current_user.id)
        elif resource == "payments":
            query = query.eq("owner_id", current_user.id)

    if status_filter:
        query = query.eq("status", status_filter)
    if state_filter and resource == "properties":
        query = query.ilike("state", f"%{state_filter}%")
    if property_type_filter and resource == "properties":
        query = query.eq("property_type", property_type_filter)

    query = query.order("created_at", desc=True).range(skip, skip + limit - 1)
    return query.execute()


def _make_csv(rows: list[dict], columns: list[str]) -> io.StringIO:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow([str(row.get(c, "") or "") for c in columns])
    buf.seek(0)
    return buf


def _make_xlsx(rows: list[dict], columns: list[str]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    header_font = Font(bold=True)
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = header_font

    for r, row in enumerate(rows, 2):
        for c, col in enumerate(columns, 1):
            ws.cell(row=r, column=c, value=row.get(col, "") or "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/properties")
def export_properties(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    status: str | None = Query(None),
    state: str | None = Query(None),
    property_type: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=50000),
    current_user: CurrentUser = Depends(require_super_admin_or_manager),
    supabase: Client = Depends(get_service_client),
):
    result = _build_query(supabase, "properties", current_user, status, state, property_type, skip, limit)
    rows = result.data or []
    ext = format
    media = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf = _make_csv(rows, COLUMNS["properties"]) if format == "csv" else _make_xlsx(rows, COLUMNS["properties"])
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f"attachment; filename=properties_{date.today()}.{ext}"},
    )


@router.get("/tenants")
def export_tenants(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    status: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=50000),
    current_user: CurrentUser = Depends(require_super_admin_or_manager),
    supabase: Client = Depends(get_service_client),
):
    result = _build_query(supabase, "tenants", current_user, status, skip=skip, limit=limit)
    rows = result.data or []
    ext = format
    media = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf = _make_csv(rows, COLUMNS["tenants"]) if format == "csv" else _make_xlsx(rows, COLUMNS["tenants"])
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f"attachment; filename=tenants_{date.today()}.{ext}"},
    )


@router.get("/leases")
def export_leases(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    status: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=50000),
    current_user: CurrentUser = Depends(require_super_admin_or_manager),
    supabase: Client = Depends(get_service_client),
):
    result = _build_query(supabase, "leases", current_user, status, skip=skip, limit=limit)
    rows = result.data or []
    ext = format
    media = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf = _make_csv(rows, COLUMNS["leases"]) if format == "csv" else _make_xlsx(rows, COLUMNS["leases"])
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f"attachment; filename=leases_{date.today()}.{ext}"},
    )


@router.get("/payments")
def export_payments(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    status: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=50000),
    current_user: CurrentUser = Depends(require_super_admin_or_manager),
    supabase: Client = Depends(get_service_client),
):
    result = _build_query(supabase, "payments", current_user, status, skip=skip, limit=limit)
    rows = result.data or []
    ext = format
    media = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf = _make_csv(rows, COLUMNS["payments"]) if format == "csv" else _make_xlsx(rows, COLUMNS["payments"])
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f"attachment; filename=payments_{date.today()}.{ext}"},
    )


@router.get("/boosts")
def export_boosts(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    status: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=50000),
    current_user: CurrentUser = Depends(require_super_admin),
    supabase: Client = Depends(get_service_client),
):
    result = _build_query(supabase, "boosts", current_user, status, skip=skip, limit=limit)
    rows = result.data or []
    ext = format
    media = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf = _make_csv(rows, COLUMNS["boosts"]) if format == "csv" else _make_xlsx(rows, COLUMNS["boosts"])
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f"attachment; filename=boosts_{date.today()}.{ext}"},
    )


@router.get("/report-pdf")
def export_report_pdf(
    current_user: CurrentUser = Depends(require_super_admin_or_manager),
    supabase: Client = Depends(get_service_client),
):
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    props = supabase.table("properties").select("title, property_type, bedrooms, monthly_rent, status, city, state").eq("owner_id", current_user.id).execute()
    tenants_data = supabase.table("tenants").select("first_name, last_name, email, phone, status").eq("owner_id", current_user.id).execute()
    leases_data = supabase.table("leases").select("monthly_rent, status, start_date, end_date").eq("owner_id", current_user.id).execute()
    payments_data = supabase.table("payments").select("amount, status, paid_date").eq("owner_id", current_user.id).execute()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=18*mm, leftMargin=18*mm, topMargin=18*mm, bottomMargin=18*mm, title="Portfolio Report")
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("Title", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, textColor=colors.HexColor("#0F766E"), spaceAfter=12)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=13, spaceAfter=6, spaceBefore=14)
    body = styles["Normal"]

    story = [Paragraph("Afodabo Housing", styles["Normal"]), Paragraph("Portfolio Report", title_s), Spacer(1, 6*mm)]

    p_rows = [[r.get("title",""), r.get("property_type",""), str(r.get("bedrooms","")), f"UGX {r['monthly_rent']:,.0f}" if r.get("monthly_rent") else "", r.get("status",""), f"{r.get('city','')}, {r.get('state','')}"] for r in (props.data or [])]
    if p_rows:
        story.append(Paragraph(f"Properties ({len(p_rows)})", h2))
        story.append(Table([["Title","Type","Beds","Rent","Status","Location"]] + p_rows, colWidths=[60*mm,30*mm,16*mm,30*mm,22*mm,40*mm], repeatRows=1, hAlign="LEFT"))

    t_rows = [[f"{r.get('first_name','')} {r.get('last_name','')}".strip(), r.get("email",""), r.get("phone",""), r.get("status","")] for r in (tenants_data.data or [])]
    if t_rows:
        story.append(Spacer(1, 6*mm))
        story.append(Paragraph(f"Tenants ({len(t_rows)})", h2))
        story.append(Table([["Name","Email","Phone","Status"]] + t_rows, colWidths=[40*mm,60*mm,50*mm,28*mm], repeatRows=1, hAlign="LEFT"))

    total_rent = sum(float(r.get("monthly_rent", 0)) for r in (leases_data.data or []))
    total_paid = sum(float(r.get("amount", 0)) for r in (payments_data.data or []) if r.get("status") in ("confirmed","completed"))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph("Summary", h2))
    story.append(Table([["Metric","Value"],[
        "Total Monthly Rent", f"UGX {total_rent:,.0f}"],
        ["Total Collected", f"UGX {total_paid:,.0f}"],
        ["Active Leases", str(sum(1 for l in (leases_data.data or []) if l.get("status")=="active"))],
        ["Tenants", str(len(t_rows))],
    ], colWidths=[60*mm, 60*mm], hAlign="LEFT"))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(f"Generated on {date.today().isoformat()}", body))
    doc.build(story)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="portfolio_report_{date.today()}.pdf"'},
    )


@router.get("/managers")
def export_managers(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    status: str | None = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=50000),
    current_user: CurrentUser = Depends(require_super_admin),
    supabase: Client = Depends(get_service_client),
):
    result = _build_query(supabase, "managers", current_user, status, skip=skip, limit=limit)
    rows = result.data or []
    ext = format
    media = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    buf = _make_csv(rows, COLUMNS["managers"]) if format == "csv" else _make_xlsx(rows, COLUMNS["managers"])
    return StreamingResponse(
        buf, media_type=media,
        headers={"Content-Disposition": f"attachment; filename=managers_{date.today()}.{ext}"},
    )
